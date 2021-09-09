import sys
from PyQt5 import QtWidgets
import requests
import json
from openpyxl import Workbook,load_workbook
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import *
from DepoAnaEkranUI import *
from hakkında import *
from ÜrünEkle import *
from Döviz import *
from VeriAl import * 
from guncelle import *

#----------------------OYGULAMA OLUŞTURMA----------------------#
#----------------------OYGULAMA OLUŞTURMA----------------------#
Uygulama = QtWidgets.QApplication(sys.argv)
penAna = QMainWindow()
ui = Ui_MainWindow()
ui.setupUi(penAna)
penAna.show()


penHakında = QDialog()
ui2 = Ui_Dialog()
ui2.setupUi(penHakında)

penEkle = QWidget()
uiEkle = Ui_FormE()
uiEkle.setupUi(penEkle)

penDöviz = QWidget()
uiDöviz = Ui_Form()
uiDöviz.setupUi(penDöviz)

penveri = QWidget()
uiVeri = Ui_FormV()
uiVeri.setupUi(penveri)

penGüncelle = QWidget()
uiGüncelle = Ui_FormG()
uiGüncelle.setupUi(penGüncelle)

#----------------------VERİTABANI OLUŞTUR-----------------#
#----------------------VERİTABANI OLUŞTUR-----------------#


import sqlite3
global curs
global conn 
conn = sqlite3.connect("DepoVeritabani.db")
curs = conn.cursor()
sorguCreTBLDepo = ("CREATE TABLE IF NOT EXISTS depo(             \
                  ürünAdı TEXT ,                         \
                  DepoKodu NOT NULL PRIMARY KEY ,                  \
                  Açıklama TEXT,     \
                  Marka TEXT,   \
                  DolapNo TEXT, \
                  RafNo TEXT,  \
                  ÜrünAdedi TEXT, \
                  KritikStok TEXT)")
                  

curs.execute(sorguCreTBLDepo)
conn.commit()



#----------------------Listele-----------------------------#
#----------------------listele-----------------------------#

def LISTELE():
    try:
        ui.urunTablosu.clear()

        ui.urunTablosu.setHorizontalHeaderLabels(("Ürün Adı","Depo Kodu","Açıklama","Marka","Dolap No","Raf No","Ürün Adedi","Kritik Stok"))

        ui.urunTablosu.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        curs.execute("SELECT * FROM depo")

        for satirIndeks, satirVeri in enumerate(curs):
            for sutunIndeks, sutunVeri in enumerate (satirVeri):
                ui.urunTablosu.setItem(satirIndeks,sutunIndeks,QTableWidgetItem(str(sutunVeri)))

        uiEkle.ln_urunAdi.clear()
        uiEkle.ln_urunAciklamasi.clear()
        uiEkle.ln_markaEkle.clear()
        uiEkle.ln_urunAdedi.clear()
        uiEkle.ln_RafNo.clear()
        uiEkle.ln_dolapNo.clear()
        uiEkle.ln_depoKodu.clear()
        uiEkle.lineEdit.clear()
    except Exception as hata_Lıstele:
        ui.statusbar.showMessage("Hata_Lıstele: "+str(hata_Lıstele),6000)
    

LISTELE()


#----------------------EKEL-----------------------------#
#----------------------EKLE-----------------------------#
def EKLE():
    try:
        _lnUrunAdi = uiEkle.ln_urunAdi.text()
        _lnUrunhakkinda = uiEkle.ln_urunAciklamasi.text()
        _lnUrunAdedi = uiEkle.ln_urunAdedi.text()
        _lndepoKodu = uiEkle.ln_depoKodu.text()
        _lnRafNo = uiEkle.ln_RafNo.text()
        _lndolapNo = uiEkle.ln_dolapNo.text()
        _marka = uiEkle.ln_markaEkle.text()
        _kritikStok = uiEkle.lineEdit.text()

    
        if _lndepoKodu == "":
            eksik_bilgi = QMessageBox()
            eksik_bilgi.setIcon(QMessageBox.Information)
            eksik_bilgi.setText("Lütfen tüm bilgileri doldurun")
            eksik_bilgi.setWindowTitle("Eksik Bilgi")
            eksik_bilgi.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)               
            returnValue = eksik_bilgi.exec()
        else:
            curs.execute("INSERT INTO depo \
                            (ürünAdı,DepoKodu,Açıklama,Marka,DolapNo,RafNo,ÜrünAdedi,KritikStok)\
                            VALUES (?,?,?,?,?,?,?,?)", \
                            (_lnUrunAdi,_lndepoKodu,_lnUrunhakkinda,_marka,_lndolapNo,_lnRafNo,_lnUrunAdedi,_kritikStok))
    except Exception as hata_ekele:
        ui.statusbar.showMessage("Hata_Ekle: "+str(hata_ekele),6000)
    conn.commit()

    LISTELE()

#----------------------Sorgu-----------------------------#
#----------------------Sorgu-----------------------------#
def SORGU ():
    try:
        arananAD = ui.ln_ekleUruAdi.text()
        arananDepoKodu = ui.ln_EkleDepoKodu.text()
        arananAçıklama = ui.ln_EkleAcklama.text()
        arananMarka = ui.ln_marka.text()

    
        curs.execute("SELECT * FROM depo WHERE ürünAdı=? OR  DepoKodu=? OR Açıklama=? OR Marka=?",\
                (arananAD,arananDepoKodu,arananAçıklama,arananMarka))
        conn.commit()
        ui.urunTablosu.clear()
        ui.urunTablosu.setHorizontalHeaderLabels(("Ürün Adı","Depo Kodu","Açıklama","Marka","Dolap No","Raf No","Ürün Adedi","Kritik Stok"))
        for satirIndeks, satirVeri in enumerate(curs):
            for sutunIndeks, sutunVeri in enumerate (satirVeri):
                ui.urunTablosu.setItem(satirIndeks,sutunIndeks,QTableWidgetItem(str(sutunVeri)))
    except Exception as hata_sorgu:
        ui.statusbar.showMessage("Hata_Sorgu: "+str(hata_sorgu),6000)

    ui.ln_ekleUruAdi.clear()
    ui.ln_EkleDepoKodu.clear()
    ui.ln_EkleAcklama.clear()
    ui.ln_marka.clear()
#----------------------Arttır-----------------------------#
#----------------------Arttır-----------------------------#
    
def StokArttır():
    try:
        secili = ui.urunTablosu.selectedItems()

        aded = secili[6].text()
        Depo_No = secili[1].text()
        kritik_stok = secili[7].text()

        ui.azaltDepoNO.setText(str(Depo_No))

        eklenecek_aded = ui.ln_SilAdet.text()

        yeniAded = int(aded) + int(eklenecek_aded)
        yeniAded = str(yeniAded)
        
        curs.execute("UPDATE depo SET ÜrünAdedi=? WHERE DepoKodu=?",\
            (yeniAded,Depo_No))
        conn.commit()
        LISTELE()
    except Exception as hata_Arttır:
        ui.statusbar.showMessage("Hata_Arttır: "+str(hata_Arttır),6000)

    



    

#----------------------Azalt-----------------------------#
#----------------------Azalt-----------------------------#    
def StokAzalt():
    try:
        secili = ui.urunTablosu.selectedItems()

        aded = secili[6].text()
        Depo_No = secili[1].text()
        kritik_stok = secili[7].text()

        ui.azaltDepoNO.setText(str(Depo_No))
        eklenecek_aded = ui.ln_SilAdet.text()
        yeniAded = int(aded) - int(eklenecek_aded)
        yeniAded = str(yeniAded)

        curs.execute("UPDATE depo SET ÜrünAdedi=? WHERE DepoKodu=?",\
            (yeniAded,Depo_No))
        conn.commit()
        LISTELE()

        if int(kritik_stok)>=int(yeniAded):

            uyarı = QMessageBox()
            uyarı.setIcon(QMessageBox.Warning)
            uyarı.setText("depo kodu: "+Depo_No+" olan ürün kritik stoğun altına indi !!!")
            uyarı.setWindowTitle("Uyarı'Kritik Değer'")
            uyarı.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)               
            returnValue = uyarı.exec()
            if returnValue==QMessageBox.Ok:
                penAna.show()
            else:
                penAna.show()
            
        
    except Exception as hata_azalt:
        ui.statusbar.showMessage("Hata_Azalt: "+str(hata_azalt),6000)
    

#----------------------yazdır-----------------------------#
#----------------------yazdır-----------------------------# 
def yazdır():
    try:
        seçili = ui.urunTablosu.selectedItems()
        uiGüncelle.ln_urunAdi_guncel.setText(seçili[0].text())
        uiGüncelle.ln_depoKodu_guncel.setText(seçili[1].text())
        uiGüncelle.ln_urunAciklamasi_guncel.setText(seçili[2].text())
        uiGüncelle.ln_markaEkle_guncel.setText(seçili[3].text())
        uiGüncelle.ln_dolapNo_guncel.setText(seçili[7].text())
        uiGüncelle.ln_RafNo_Guncel.setText(seçili[5].text())
    except Exception as hata_Yazdır:
        ui.statusbar.showMessage("Hata_Yazdır: "+str(hata_Yazdır),6000)
    
    

#----------------------Güncelle-----------------------------#
#----------------------Güncelle-----------------------------#
def guncelle():
    try:
        secili = ui.urunTablosu.selectedItems()
        depo_no_eski = secili[1].text()


        değişecek_urun_adi = uiGüncelle.ln_urunAdi_guncel.text()
        değişecek_depo_kodu = uiGüncelle.ln_depoKodu_guncel.text()
        değişecek_urun_acıklamasi = uiGüncelle.ln_urunAciklamasi_guncel.text()
        değişecek_dolap_no = uiGüncelle.ln_dolapNo_guncel.text()
        değişecek_raf_no = uiGüncelle.ln_RafNo_Guncel.text()
        değişecek_marka = uiGüncelle.ln_markaEkle_guncel.text()


        curs.execute("UPDATE depo SET ürünAdı=?, DepoKodu=?, Açıklama=? ,Marka=? ,DolapNo=? ,RafNo=?  WHERE DepoKodu=?",\
                (değişecek_urun_adi, değişecek_depo_kodu, değişecek_urun_acıklamasi, değişecek_marka, değişecek_dolap_no, değişecek_raf_no , depo_no_eski))
        conn.commit()
        LISTELE()
        
        uiGüncelle.ln_urunAdi_guncel.clear()
        uiGüncelle.ln_depoKodu_guncel.clear()
        uiGüncelle.ln_urunAciklamasi_guncel.clear()
        uiGüncelle.ln_dolapNo_guncel.clear()
        uiGüncelle.ln_RafNo_Guncel.clear()
        uiGüncelle.ln_markaEkle_guncel.clear()

    except Exception as hata_Guncelle:
        ui.statusbar.showMessage("Hata_Güncelle: "+str(hata_Guncelle),6000)
    



#----------------------VeriAl-----------------------------#
#----------------------VeriAl-----------------------------#    

def VeriAl():
    try:
        wb = load_workbook("TSY_Depo.xlsx")
        ws = wb.sheetnames
        ws = wb["Depo"]


        list_satir = []
        satır_1 = 1
        satır_2 = 2
        Açıklama = "Açıklama Yok"


        while 1508>=satır_1:
            satır_1 +=1
            satır_2 +=1
            for satir in range(satır_1,satır_2):
                for sutun in range(1,10):
                    list_satir.append(str(ws.cell(satir,sutun).value))
            
            curs.execute("INSERT INTO depo \
                                    (ürünAdı,DepoKodu,Açıklama,Marka,DolapNo,RafNo,ÜrünAdedi,KritikStok)\
                                    VALUES (?,?,?,?,?,?,?,?)", \
                                    (list_satir[1],list_satir[3],Açıklama,list_satir[2],list_satir[4],list_satir[5],list_satir[8],list_satir[7]))
            conn.commit()
            list_satir.clear()
    except Exception as hata_veri_al:
        ui.statusbar.showMessage("Hata: "+str(hata_veri_al),7000)
    

    LISTELE()

#----------------------ÜrünüKaldır-----------------------------#
#----------------------ÜrünüKladır-----------------------------#
def urunKadlır():
    try:
        cevap_kaldır = QMessageBox.question(penAna,"Ürün Kaldır!!","Ürünü istediğinize emin misiniz?",\
                            QMessageBox.Yes | QMessageBox.No)
        if cevap_kaldır == QMessageBox.Yes:
            seçili_kaldır = ui.urunTablosu.selectedItems()
            seçili_kaldır_depo_no = seçili_kaldır[1].text()
            curs.execute("DELETE FROM depo WHERE DepoKodu='%s'" %(seçili_kaldır_depo_no))
            conn.commit()
            LISTELE()
            ui.statusbar.showMessage("Ürün kaldırıldı",6000)
        else:
            ui.statusbar.showMessage("Kadırma işlemi iptal edildi",6000)
    except Exception as hata_kaldır:
        ui.statusbar.showMessage("Hata: "+str(hata_kaldır),1000)
    

#----------------------KritikStokYenile-----------------------------#
#----------------------KritikSıtokYenile-----------------------------#

def kritik_satok_yenile():
    try:
        secili_kritik = ui.urunTablosu.selectedItems()
        Depo_No = secili_kritik[1].text()
        aded_kritik = secili_kritik[6].text()
        yeni_kıritik = ui.Kritk_stok_degistir.text()

        ui.Krtik_satok_No.setText(str(Depo_No))

        curs.execute("UPDATE depo SET KritikStok=? WHERE DepoKodu=?",\
                (yeni_kıritik,Depo_No))
        conn.commit()
        LISTELE()


        if int(aded_kritik)<=int(yeni_kıritik):

                uyarı = QMessageBox()
                uyarı.setIcon(QMessageBox.Warning)
                uyarı.setText("depo kodu: "+Depo_No+" olan ürün kritik stoğun altına indi !!!")
                uyarı.setWindowTitle("Uyarı'Kritik Değer'")
                uyarı.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)               
                returnValue = uyarı.exec()
                if returnValue==QMessageBox.Ok:
                    penAna.show()
                else:
                    penAna.show()
    except Exception as hata_kritik:
        ui.statusbar.showMessage("Hata: "+str(hata_kritik),1000)

#----------------------DÖVİZ-----------------------------#
#----------------------DÖVİZ-----------------------------#
def doviz():
    try:
        api_url_dolar = "https://finans.truncgil.com/today.json"
        Dolar = requests.get(api_url_dolar)
        Dolar = json.loads(Dolar.text)
        Yazacak_dolar = str(Dolar["ABD DOLARI"]["Alış"])
        uiDöviz.lb_dolar.setText(Yazacak_dolar[0:4])


        api_url_euro = "https://finans.truncgil.com/today.json"
        Euro = requests.get(api_url_euro)
        Euro = json.loads(Euro.text)
        Yazılacak_Euro = str(Euro["EURO"]["Alış"])
        uiDöviz.lb_euro_2.setText(Yazılacak_Euro[0:4])

        api_url_kdoar = "https://finans.truncgil.com/today.json"
        kdolar = requests.get(api_url_kdoar)
        kdolar = json.loads(kdolar.text)
        Yazılacak_kdolar = str(kdolar["KANADA DOLARI"]["Alış"])
        uiDöviz.lb_kdolar.setText(Yazılacak_kdolar[0:4])

        api_url_yuan = "https://finans.truncgil.com/today.json"
        Yuan = requests.get(api_url_yuan)
        Yuan = json.loads(Yuan.text)
        Yazılacak_Yuan = str(Yuan["ÇİN YUANI"]["Alış"])
        uiDöviz.lb_yuan.setText(Yazılacak_Yuan[0:4])


        api_url_ruble = "https://finans.truncgil.com/today.json"
        Ruble = requests.get(api_url_ruble)
        Ruble = json.loads(Ruble.text)
        Yazılacak_Ruble = str(Ruble["RUS RUBLESİ"]["Alış"]) 
        uiDöviz.lb_ruble.setText(Yazılacak_Ruble[0:4])

        api_url_sterlin = "https://finans.truncgil.com/today.json"
        Sterlin = requests.get(api_url_sterlin)
        Sterlin = json.loads(Sterlin.text)
        Yazılacak_Sterlin = str(Sterlin["İNGİLİZ STERLİNİ"]["Alış"]) 
        uiDöviz.lb_sterlin.setText(Yazılacak_Sterlin[0:4])
    except Exception as hata_döviz:
        ui.statusbar.showMessage("Hata_Doviz: "+str(hata_döviz),6000)
            






#----------------------çıkış-----------------------------#
#----------------------Cıkış-----------------------------#
def CIKIS():
    cevap = QMessageBox.question(penAna,"ÇIKIŞ","Çıkmak isteiğinize eminmisiniz?",\
                            QMessageBox.Yes | QMessageBox.No)
    
    if cevap==QMessageBox.Yes:
        conn.close()
        sys.exit(Uygulama.exec_())
    else:
        penAna.show()   
#-----------------------GüncelleAç----------------------------#
#----------------------GüncelleAç-----------------------------#
def guncelleAç():
    penGüncelle.show()
#----------------------EkleAç-----------------------------#
#----------------------EkleAç-----------------------------#
def EkleAç():
    penEkle.show()

#----------------------DövizAç-----------------------------#
#----------------------DövizAç-----------------------------#
def DövizAç():
    penDöviz.show()


#----------------------Hakkında-----------------------------#
#----------------------Hakkında-----------------------------#
def HAKKINDA():
    penHakında.show()

#----------------------VeriAlAç-----------------------------#
#----------------------VeriAlAç-----------------------------#
def VeriAç():
    penveri.show()

#----------------------SİNYAL-SLOT-----------------------------#
#----------------------SİNYAL-SLOT-----------------------------#
ui.bt_listele.clicked.connect(LISTELE)
ui.bt_cikis.clicked.connect(CIKIS)
ui.bt_UrunEkleGit.clicked.connect(EkleAç)
uiEkle.bt_ekle.clicked.connect(EKLE)
ui.menuHakkinde.triggered.connect(HAKKINDA)
ui.bt_sorgula.clicked.connect(SORGU)
ui.bt_Doviz.clicked.connect(DövizAç)
ui.pb_Stokarttir.clicked.connect(StokArttır)
ui.bt_kaldr.clicked.connect(StokAzalt)
ui.menuExcel.triggered.connect(VeriAç)
uiVeri.pb_veri.clicked.connect(VeriAl)
ui.Kritk_stok_degistir_bt.clicked.connect(kritik_satok_yenile)
ui.guncelle.clicked.connect(guncelleAç)
uiGüncelle.Doldur.clicked.connect(yazdır)
uiGüncelle.bt_guncelle.clicked.connect(guncelle)
ui.Kaldr_bt.clicked.connect(urunKadlır)
ui.bt_Doviz.clicked.connect(doviz)




sys.exit(Uygulama.exec_())